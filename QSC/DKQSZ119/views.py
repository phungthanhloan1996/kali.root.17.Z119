from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from .forms import WorkerInfoForm
from django.contrib import messages
from django.http import JsonResponse
from datetime import date
class CustomLoginForm(AuthenticationForm):
    pass
def custom_login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if user.is_leader:
                return redirect('dashboard')
            elif user.is_director:
                return redirect('director_dashboard')
            else:
                return redirect('dashboard')
    else:
        form = AuthenticationForm()
    return render(request, "login.html", {"form": form})
@login_required
def dashboard_view(request):
    user = request.user
    full_name = user.get_full_name()
    today = timezone.localtime(timezone.now()).date()
    if not user.is_leader:
        return render(request, 'director_dashboard.html', {'workers': []})
    workers = WorkerInfo.objects.filter(z119group=user.z119group)
    attendance_records = AttendanceRecord.objects.filter(
        worker__in=workers, date=today
    ).order_by('worker', '-date')
    attendance_dict = {}
    for record in attendance_records:
        if record.worker.id not in attendance_dict:
            attendance_dict[record.worker.id] = record
    status_counter = {
        'present': 0,
        'presenttg': 0,
        'curve': 0,
        'absent': 0,
        'leave': 0,
        'pregnancy': 0,
        'gostation': 0,
        'study': 0,
        'studynuangayinZ':0,
        'studyfullinZ':0,
        'nghichohuu':0,
        'kinhte':0,
    }
    for record in attendance_dict.values():
        if record.status in status_counter:
            status_counter[record.status] += 1
    if request.method == 'POST':
        updated = False
        now = timezone.localtime(timezone.now()) 
        for worker in workers:
            status = request.POST.get(f'status_{worker.id}')
            note = request.POST.get(f'note_{worker.id}', '')
            record = attendance_dict.get(worker.id)
            if status:
                if record:
                    record.status = status
                    record.note = note
                    record.report_time = now
                    record.save()
                else:
                    AttendanceRecord.objects.create(
                        worker=worker,
                        date=today,
                        status=status,
                        note=note,
                        report_time=now
                    )
                updated = True
        if updated:
            messages.success(request, "Đã lưu chấm công cho hôm nay.")
        else:
            messages.error(request, "Bạn chưa chọn trạng thái nào để lưu.")
        return redirect('dashboard')
    for worker in workers:
        record = attendance_dict.get(worker.id)
        total_worker = len(workers)
        setattr(worker, 'attendance_today', record)
        present_count = status_counter['present']+status_counter['presenttg']+status_counter['gostation']+status_counter['studynuangayinZ']+status_counter['studyfullinZ']
        absent_statuses = status_counter['curve']+status_counter['absent']+status_counter['leave']+status_counter['pregnancy']+status_counter['study']+status_counter['nghichohuu']+status_counter['kinhte']
    return render(request, 'dashboard.html', {
        'workers': workers,
        'today': today,
        'leader_group': user.z119group.name,
        'leader_name': full_name,
        'total_present': status_counter['present'],
        'total_presenttg': status_counter['presenttg'],
        'total_curve': status_counter['curve'],
        'total_absent': status_counter['absent'],
        'total_leave': status_counter['leave'],
        'total_gostation': status_counter['gostation'],
        'total_pregnancy': status_counter['pregnancy'],
        'total_study': status_counter['study'],
        'total_studynuangayinZ':status_counter['studynuangayinZ'],
        'total_studyfullinZ':status_counter['studyfullinZ'],
        'total_nghichohuu':status_counter['nghichohuu'],
        'total_kinhte':status_counter['kinhte'],
        'total_worker':total_worker,
        'present_count':present_count,
        'absent_statuses':absent_statuses,
    })
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .models import Z119group, CustomUser
@login_required
def chamcongngay(request):
    today = timezone.localtime(timezone.now()).date()
    weekday_map = {
        0: "Thứ 2",
        1: "Thứ 3",
        2: "Thứ 4",
        3: "Thứ 5",
        4: "Thứ 6",
        5: "Thứ 7",
        6: "Chủ nhật"
    }
    weekday_str = weekday_map[today.weekday()]
    day = today.day
    month = today.month
    year = today.year
    formatted_date = f"{weekday_str} ngày {day} tháng {month} năm {year}"

    groups = Z119group.objects.all()
    report_data = []

    for group in groups:
        workers = WorkerInfo.objects.filter(z119group=group)
        leave_types = {
            'leave': [],
            'curve': [],
            'pregnancy': [],
            'nghichohuu': [],
            'study': [],
            'absent':[],
            'kinhte':[],
            'khac': [],
        }
        absent_count = 0

        for worker in workers:
            attendance_record = AttendanceRecord.objects.filter(worker=worker, date=today).first()
            if attendance_record:
                if attendance_record.status in leave_types:
                    leave_types[attendance_record.status].append(worker.name)
                    absent_count += 1
                    if not attendance_record.report_time:
                        attendance_record.report_time = timezone.localtime(timezone.now())
                        attendance_record.save()
        group_records = AttendanceRecord.objects.filter(
            worker__in=workers, date=today, report_time__isnull=False
        ).order_by('report_time')
        report_time = group_records.first().report_time if group_records.exists() else None
        leader = CustomUser.objects.filter(z119group=group, is_leader=True).first()
        reporter_name = leader.get_full_name() if leader else "Chưa có người báo cáo"
        working_count = workers.count() - absent_count
        note = f"{workers.count()} / {working_count} / {absent_count}"
        report_data.append({
            'group': group.name,
            'workers': workers,
            'report_time': report_time,
            'leave_types': leave_types,
            'absent_count': absent_count,
            'working_count': working_count,
            'reporter': reporter_name,
            'note': note,
        })
    return render(request, 'chamcongngay.html', {
        'today_full': formatted_date,
        'report_data': report_data,
    })
@login_required
def add_worker_view(request):
    if request.method == 'POST':
        form = WorkerInfoForm(request.POST)
        if form.is_valid():
            worker = form.save(commit=False)
            worker.z119group = request.user.z119group
            worker.save()
            today = timezone.now().date()
            AttendanceRecord.objects.create(
                worker=worker,
                date=today,
                status='absent',
                note=''
            )
            return redirect('dashboard')
    else:
        form = WorkerInfoForm()
    return render(request, 'add_worker.html', {'form': form})
@login_required
def edit_all_worker(request):
    workers = WorkerInfo.objects.filter(z119group=request.user.z119group)
    if request.method == 'POST':
        for worker in workers:
            name = request.POST.get(f'name_{worker.id}')
            rank = request.POST.get(f'rank_{worker.id}')
            position = request.POST.get(f'position_{worker.id}')
            phone_number = request.POST.get(f'phone_number_{worker.id}')
            if name: worker.name = name
            if rank: worker.rank = rank
            if position: worker.position = position
            if phone_number: worker.phone_number = phone_number
            worker.save()
        messages.success(request, "Cập nhật thông tin công nhân thành công.")
        return redirect('edit_all_worker')

    return render(request, 'edit_all_worker.html', {'workers': workers})
@login_required
def delete_worker(request):
    if request.method == 'POST':
        worker_id = request.POST.get('worker_id')
        if worker_id:
            worker = get_object_or_404(WorkerInfo, id=worker_id)
            if request.user.is_leader and worker.z119group == request.user.z119group:
                worker.delete()
                messages.success(request, f'Công nhân {worker.name} đã được xóa thành công!')
            else:
                messages.error(request, 'Bạn không có quyền xóa công nhân này.')
        else:
            messages.error(request, 'Có lỗi xảy ra khi xóa công nhân.')
        return redirect('delete_worker')
    else:
        if request.user.is_leader:
            workers = WorkerInfo.objects.filter(z119group=request.user.z119group)
        else:
            workers = WorkerInfo.objects.none()
        return render(request, 'delete_worker.html', {'workers': workers})
from django.utils.timezone import now
from django.db.models import Q
from .models import Z119group, AttendanceRecord, WorkerInfo
def get_Z119group_not_reported_today():
    today = timezone.now().date()
    group_with_workers = Z119group.objects.filter(workerinfo__isnull=False).distinct()
    group_reported = Z119group.objects.filter(workerinfo__attendancerecord__date=today).distinct()
    group_not_reported = group_with_workers.exclude(id__in=group_reported)
    return group_not_reported
@login_required
def director_dashboard(request):
    today = date.today()
    total_workers = WorkerInfo.objects.count()
    attendance_today = AttendanceRecord.objects.filter(date=today)
    present_statuses = ['present', 'presenttg','gostation','studynuangayinZ','studyfullinZ']
    leave_statuses = ['leave']
    study_statuses = ['study']
    pregnancy_statuses = ['pregnancy']
    nghichohuu_statuses = ['nghichohuu']
    curve_statuses = ['curve']
    traiphep_statuses = ['absent']
    kinhte_statuses = ['kinhte']
    leave_count = attendance_today.filter(status__in=leave_statuses).count()
    study_count = attendance_today.filter(status__in=study_statuses).count()
    pregnancy_count = attendance_today.filter(status__in=pregnancy_statuses).count()
    nghichohuu_count = attendance_today.filter(status__in=nghichohuu_statuses).count()
    present_count = attendance_today.filter(status__in=present_statuses).count()
    curve_count = attendance_today.filter(status__in=curve_statuses).count()
    absent_records = attendance_today.exclude(status__in=present_statuses)
    traiphep_count = attendance_today.filter(status__in=traiphep_statuses).count()
    kinhte_count = attendance_today.filter(status__in=kinhte_statuses).count()
    absent_count = absent_records.count()
    not_reported_groups = get_Z119group_not_reported_today()
    context = {
        'total_workers': total_workers,
        'present_count': present_count,
        'absent_count': absent_count,
        'absent_records': absent_records,
        'today': today,
        'leave_count':leave_count,
        'study_count' :study_count,
        'pregnancy_count': pregnancy_count,
        'nghichohuu_count': nghichohuu_count,
        'traiphep_count': traiphep_count,
        'curve_count': curve_count,
        'kinhte_count':kinhte_count,
        'not_reported_groups':not_reported_groups,
    }
    return render(request, 'director_dashboard.html', context)

@login_required
def get_realtime_attendance(request):
    today = timezone.now().date()
    records = AttendanceRecord.objects.filter(date=today)
    data = []
    for r in records:
        data.append({
            'name': r.worker.name,
            'group': r.worker.z119group.name,
            'rank': r.worker.rank,
            'position': r.worker.position,
            'status': r.get_status_display(),
            'note': r.note or ''
        })
    return JsonResponse({'data': data})
@login_required
def logout_view(request):
    logout(request)
    return redirect('login')






    
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from datetime import datetime
import calendar
from .models import WorkerInfo, AttendanceRecord
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from datetime import datetime
import calendar
from django.http import HttpResponse

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO

from .models import WorkerInfo, AttendanceRecord

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime, date, timedelta
import calendar
from .models import WorkerInfo, AttendanceRecord

STATUS_MAP = {
    'present': 'SP',
    'presenttg': '+',
    'curve': 'CT',
    'absent': 'V',
    'leave': 'P',
    'gostation': 'G',
    'pregnancy': 'TS',
    'study': 'H',
    'studynuangayinZ': 'H/SP',
    'studyfullinZ': 'H_i_Z',
    'nghichohuu': 'CH',
    'kinhte': 'KT',
}

@login_required
def chamcong(request):
    user = request.user
    now = timezone.localtime(timezone.now())

    # --- Lấy tháng đang chọn (GET param month=YYYY-MM) ---
    month_str = request.GET.get("month")
    if month_str:
        try:
            if "-" in month_str:
                selected_date = datetime.strptime(month_str, "%Y-%m").date()
            else:
                selected_date = date(now.year, int(month_str), 1)
        except ValueError:
            selected_date = date(now.year, now.month, 1)
    else:
        selected_date = date(now.year, now.month, 1)

    month = selected_date.month
    year = selected_date.year

    # --- Tính tháng trước / sau ---
    prev_month_date = (selected_date.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_month_date = (selected_date.replace(day=28) + timedelta(days=4)).replace(day=1)

    prev_month = prev_month_date.strftime("%Y-%m")
    next_month = next_month_date.strftime("%Y-%m")

    # --- Danh sách ngày trong tháng ---
    num_days = calendar.monthrange(year, month)[1]
    days = list(range(1, num_days + 1))

    # --- Lấy danh sách nhân sự ---
    if hasattr(user, "is_director") and user.is_director:
        workers = WorkerInfo.objects.all().order_by("name")
    else:
        workers = WorkerInfo.objects.filter(z119group=user.z119group).order_by("name")

    # --- Lấy dữ liệu chấm công ---
    records = AttendanceRecord.objects.filter(
        worker__in=workers, date__year=year, date__month=month
    )

    attendance_by_worker = {w.id: {d: "" for d in days} for w in workers}

    for r in records:
        wk_id = r.worker_id
        day = r.date.day
        symbol = STATUS_MAP.get(r.status, r.status or "")
        if wk_id in attendance_by_worker and day in attendance_by_worker[wk_id]:
            attendance_by_worker[wk_id][day] = symbol

    # --- Tổng hợp ---
    users = []
    for w in workers:
        daily = attendance_by_worker[w.id]
        total_present = sum(1 for v in daily.values() if v == STATUS_MAP["present"])
        total_presenttg = sum(1 for v in daily.values() if v == STATUS_MAP["presenttg"])
        total_curve = sum(1 for v in daily.values() if v == STATUS_MAP["curve"])
        total_leave = sum(1 for v in daily.values() if v == STATUS_MAP["leave"])
        total_pregnancy = sum(1 for v in daily.values() if v == STATUS_MAP["pregnancy"])
        total_meal = (
            total_present + total_presenttg + total_curve
        )
        total_other = sum(
            1 for v in daily.values() if v not in STATUS_MAP.values() and v != ""
        )

        users.append({
            "worker": w,
            "attendance": daily,
            "total_present": total_present,
            "total_presenttg": total_presenttg,
            "total_curve": total_curve,
            "total_leave": total_leave,
            "total_pregnancy": total_pregnancy,
            "total_meal": total_meal,
            "total_other": total_other,
        })

    context = {
        "unit_name": getattr(user, "z119group").name if getattr(user, "z119group", None) else "",
        "days": days,
        "users": users,
        "month": month,
        "year": year,
        "prev_month": prev_month,
        "next_month": next_month,
    }

    return render(request, "chamcong.html", context)

@login_required
def export_attendance_excel(request):
    import calendar
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.worksheet.page import PageMargins

    user = request.user
    now = timezone.localtime(timezone.now())
    try:
        month = int(request.GET.get('month', now.month))
    except (TypeError, ValueError):
        month = now.month
    try:
        year = int(request.GET.get('year', now.year))
    except (TypeError, ValueError):
        year = now.year

    num_days = calendar.monthrange(year, month)[1]
    days = list(range(1, num_days + 1))

    if user.is_director:
        workers = list(WorkerInfo.objects.all().order_by('name'))
    else:
        workers = list(WorkerInfo.objects.filter(z119group=user.z119group).order_by('name'))

    records_qs = AttendanceRecord.objects.filter(
        worker__in=workers, date__year=year, date__month=month
    )

    attendance_by_worker = {w.id: {d: '' for d in days} for w in workers}
    for r in records_qs:
        wk_id = r.worker_id
        d = r.date.day
        attendance_by_worker[wk_id][d] = STATUS_MAP.get(r.status, r.status or '')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ChamCong_{month}_{year}"

    font = Font(name="Times New Roman", size=11)
    bold_font = Font(name="Times New Roman", size=12, bold=True)
    title_font = Font(name="Times New Roman", size=14, bold=True)
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)

    # --- TIÊU ĐỀ ĐẦU TRANG ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
    ws.cell(row=1, column=1, value="NHÀ MÁY Z119 / QUÂN CHỦNG PHÒNG KHÔNG - KHÔNG QUÂN").font = bold_font
    ws.cell(row=1, column=1).alignment = align_center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)
    ws.cell(row=2, column=1, value=f"Đơn vị: {user.z119group.name if hasattr(user, 'z119group') else ''}").alignment = align_left

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=15)
    ws.cell(row=4, column=1, value=f"BẢNG CHẤM CÔNG").font = title_font
    ws.cell(row=4, column=1).alignment = align_center

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=15)
    ws.cell(row=5, column=1, value=f"(Tháng {month}/{year})").alignment = align_center

    # --- DÒNG TIÊU ĐỀ BẢNG ---
    header_row = 7
    ws.cell(row=header_row, column=1, value="STT")
    ws.cell(row=header_row, column=2, value="Họ và tên")
    ws.cell(row=header_row, column=3, value="Chức vụ")

    start_day_col = 4
    for i, d in enumerate(days, start=start_day_col):
        ws.cell(row=header_row, column=i, value=d)

    extra_headers = ["SP", "+", "Cơm", "100%", "%", "CH", "Khác"]
    for i, h in enumerate(extra_headers, start=start_day_col + len(days)):
        ws.cell(row=header_row, column=i, value=h)

    # --- GỘP HÀNG NGÀY TRONG THÁNG ---
    ws.merge_cells(start_row=6, start_column=start_day_col, end_row=6, end_column=start_day_col + len(days) - 1)
    ws.cell(row=6, column=start_day_col, value="NGÀY TRONG THÁNG").alignment = align_center
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=3)
    ws.cell(row=6, column=1, value="").alignment = align_center
    ws.merge_cells(start_row=6, start_column=start_day_col + len(days), end_row=6, end_column=start_day_col + len(days) + len(extra_headers) - 1)
    ws.cell(row=6, column=start_day_col + len(days), value="QUY RA CÔNG").alignment = align_center

    # --- DỮ LIỆU ---
    for idx, w in enumerate(workers, start=1):
        row_idx = header_row + idx
        daily = attendance_by_worker.get(w.id, {d: '' for d in days})
        total_present = sum(1 for v in daily.values() if v == "SP")
        total_presenttg = sum(1 for v in daily.values() if v == "+")
        total_curve = sum(1 for v in daily.values() if v == "CH")
        total_leave = sum(1 for v in daily.values() if v == "P")
        total_pregnancy = sum(1 for v in daily.values() if v == "TS")
        total_other = sum(1 for v in daily.values() if v not in ["SP", "+", "CH", "P", "TS", ""])

        total_meal = total_present + total_presenttg
        row = [idx, w.name, w.position or '']
        row += [daily[d] for d in days]
        row += [total_present, total_presenttg, total_meal, total_pregnancy, total_leave, total_curve, total_other]
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = font
            cell.border = border
            cell.alignment = align_left if col == 2 else align_center

    # --- CHỮ KÝ ---
    last_row = ws.max_row + 3
    ws.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=6)
    ws.cell(row=last_row, column=2, value="GIÁM ĐỐC\n(Ký, ghi rõ họ tên)").alignment = align_center
    ws.merge_cells(start_row=last_row, start_column=7, end_row=last_row, end_column=11)
    ws.cell(row=last_row, column=7, value="PHỤ TRÁCH A-B\n(Ký, ghi rõ họ tên)").alignment = align_center
    ws.merge_cells(start_row=last_row, start_column=12, end_row=last_row, end_column=17)
    ws.cell(row=last_row, column=12, value="NGƯỜI CHẤM CÔNG\n(Ký, ghi rõ họ tên)").alignment = align_center

    note_row = last_row + 3
    ws.cell(
        row=note_row,
        column=2,
        value="Ký hiệu chấm công: SP; +; P; TS; CH"
    ).alignment = align_left

    # --- ĐIỀU CHỈNH KÍCH THƯỚC CỘT ---
    widths = {1: 5, 2: 22, 3: 14}
    for i, d in enumerate(days, start=start_day_col):
        widths[i] = 3.8
    for i, h in enumerate(extra_headers, start=start_day_col + len(days)):
        widths[i] = 9
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"BangChamCong_{month}_{year}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


from django.utils import timezone
from django.shortcuts import render
def base(request):
    now = timezone.now()
    return render(request,'base.html', {'now':now})

from datetime import datetime
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import AttendanceRecord, WorkerInfo, Z119group, CustomUser

@login_required
def history_attendance(request):
    # --- Lấy ngày được chọn từ query ?date=YYYY-MM-DD ---
    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.localtime(timezone.now()).date()
    else:
        selected_date = timezone.localtime(timezone.now()).date()

    # --- Format ngày để hiển thị ---
    weekday_map = {
        0: "Thứ 2",
        1: "Thứ 3",
        2: "Thứ 4",
        3: "Thứ 5",
        4: "Thứ 6",
        5: "Thứ 7",
        6: "Chủ nhật"
    }
    weekday_str = weekday_map[selected_date.weekday()]
    today_full = f"{weekday_str} ngày {selected_date.day} tháng {selected_date.month} năm {selected_date.year}"

    # --- Lấy dữ liệu báo cáo (giống chamcongngay) ---
    groups = Z119group.objects.all()
    report_data = []

    for group in groups:
        workers = WorkerInfo.objects.filter(z119group=group)
        leave_types = {
            'study': [],
            'leave': [],
            'curve': [],
            'pregnancy': [],
            'bu': [],
            'cho': [],
            'nghichohuu': [],
            'absent': [],
            'kinhte': [],
        }

        absent_count = 0

        for worker in workers:
            attendance_record = AttendanceRecord.objects.filter(worker=worker, date=selected_date).first()
            if attendance_record:
                status = attendance_record.status
                if status in leave_types:
                    leave_types[status].append(worker.name)
                    absent_count += 1

        # Thời gian báo cáo & người báo cáo
        group_records = AttendanceRecord.objects.filter(
            worker__in=workers,
            date=selected_date,
            report_time__isnull=False
        ).order_by('report_time')

        report_time = group_records.first().report_time if group_records.exists() else None
        leader = CustomUser.objects.filter(z119group=group, is_leader=True).first()
        reporter_name = leader.get_full_name() if leader else "Chưa có người báo cáo"

        working_count = workers.count() - absent_count
        note = f"{workers.count()} / {absent_count} / {working_count}"

        report_data.append({
            'group': group.name,
            'workers': workers,
            'report_time': report_time,
            'leave_types': leave_types,
            'absent_count': absent_count,
            'working_count': working_count,
            'reporter': reporter_name,
            'note': note,
        })

    # --- Render ra history.html ---
    return render(request, 'history.html', {
        'today_full': today_full,
        'report_data': report_data,
        'selected_date': selected_date,
    })

from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.mixins import UserPassesTestMixin
class LeaderOnlyPasswordChangeView(UserPassesTestMixin, PasswordChangeView):
    template_name = 'change_password.html'
    success_url = reverse_lazy('password_change_done')
    def test_func(self):
        return self.request.user.groups.filter(name='leader').exists()
    
